#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl import load_workbook

print("=== EXCEL FILE ANALYSIS ===")
print(f"pandas version: {pd.__version__}")
print(f"openpyxl version: {openpyxl.__version__}")

# Load the Excel file
file_path = "DUMMY DATA FOR PRECISION AREAS.xlsx"
print(f"\nAnalyzing file: {file_path}")

# First, check sheets using openpyxl
wb = load_workbook(file_path, read_only=True)
print(f"\nNumber of sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")

# Analyze each sheet
for sheet_name in wb.sheetnames:
    print(f"\n{'='*50}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*50}")
    
    # Load sheet with pandas
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        print(f"Dimensions: {df.shape[0]} rows, {df.shape[1]} columns")
        print(f"Column names: {list(df.columns)}")
        
        # Data types
        print(f"\nData types:")
        for col in df.columns:
            print(f"  {col}: {df[col].dtype}")
        
        # Sample data
        print(f"\nFirst 5 rows:")
        print(df.head())
        
        # Data quality checks
        print(f"\nData Quality Assessment:")
        print(f"  Total missing values: {df.isnull().sum().sum()}")
        missing_by_col = df.isnull().sum()
        if missing_by_col.sum() > 0:
            print("  Missing values by column:")
            for col, missing in missing_by_col.items():
                if missing > 0:
                    print(f"    {col}: {missing} ({missing/len(df)*100:.1f}%)")
        
        # Duplicate rows
        duplicates = df.duplicated().sum()
        print(f"  Duplicate rows: {duplicates}")
        
        # Basic statistics for numeric columns
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            print(f"\nNumeric columns summary:")
            print(df[numeric_cols].describe())
        
        # Unique values for categorical columns (limit to reasonable number)
        categorical_cols = df.select_dtypes(include=['object']).columns
        if len(categorical_cols) > 0:
            print(f"\nCategorical columns summary:")
            for col in categorical_cols:
                unique_count = df[col].nunique()
                print(f"  {col}: {unique_count} unique values")
                if unique_count <= 10:  # Show all if reasonable number
                    print(f"    Values: {sorted(df[col].dropna().unique().tolist())}")
        
    except Exception as e:
        print(f"Error reading sheet {sheet_name}: {e}")

wb.close()
print(f"\n{'='*50}")
print("ANALYSIS COMPLETE")
print(f"{'='*50}")